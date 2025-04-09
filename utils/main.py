from datetime import datetime
from functions import connect
import sys
#this sys.path.append( '.' ) should come before the line: from myhelpers import logger
sys.path.append( '.' )
from myhelpers import logger
import logging
import openpyxl
from logging.handlers import RotatingFileHandler

print("First output from main.py file")
print("--------")
print(sys.executable)  # Shows which Python installation is running
print(sys.path)       # Shows where Python looks for modules
print("--------")

# adding additional basic log
logging.basicConfig(level=logging.DEBUG, filename='./logs/myapp.log')

# adding more logs
logger = logging.getLogger(__name__)
stream_h = logging.StreamHandler()
file_h = logging.FileHandler('./logs/file.log')
rotating_fh = RotatingFileHandler('./logs/sliced_'+datetime.now().strftime('%m-%d-%Y--%H-%M-%S')+'.log', mode="a", maxBytes= 1000, backupCount=5)

# setting log levels
# logger.setLevel(logging.WARNING) перетирає = перекриває по пріоритету (оскільки йде нижче в файлі строку logging.basicConfig(level=logging.DEBUG, filename='./logs/myapp.log')
logger.setLevel(logging.WARNING)
# stream_h.setLevel(logging.INFO) ні на що не впливає
stream_h.setLevel(logging.INFO)
# логуються дійсно лише помилки на рівні
file_h.setLevel(logging.ERROR)

# setting log formatter
formatter = logging.Formatter('%(name)s - %(levelname)s - %(message)s')
stream_h.setFormatter(formatter)
file_h.setFormatter(formatter)

# adding hendlers
logger.addHandler(stream_h)
logger.addHandler(file_h)
logger.addHandler(rotating_fh) 

for x in range(1000):
    logger.info(f'test info message {x}')

logger.warning('This is a final warning.')

try:
    inv_file = openpyxl.load_workbook("resources/inventory.xlsx")
    product_list = inv_file["Sheet1"]

    products_per_supplier = dict()
    total_value_per_supplier = dict()
    product_no_under_21 = dict()
    print(f'{product_list.max_row} rows in sheet')

    #for product_item in product_list.max_row:

    for product_row in range(2, product_list.max_row +   1):
        supplier_name = product_list.cell(product_row, 4).value
        inventory = product_list.cell(product_row, 2).value
        price = product_list.cell(product_row, 3).value
        product_no = product_list.cell(product_row, 1)
        inventory_price = product_list.cell(product_row, 5)

        #enlisting suppliers
        if supplier_name in products_per_supplier:
            #current_number_producs = products_per_supplier[supplier_name]
            current_number_producs = products_per_supplier.get(supplier_name)
            products_per_supplier[supplier_name] = current_number_producs + 1
        else:
            print("Adding new Supplier")
            products_per_supplier[supplier_name] = 1

        #calculating total value per supplier
        if supplier_name in total_value_per_supplier:
            current_total_value_per_supplier = total_value_per_supplier.get(supplier_name) 
            total_value_per_supplier[supplier_name] = current_total_value_per_supplier + (inventory * price)
        else:
            print("Calculating first product total value per supplier")
            total_value_per_supplier[supplier_name] = inventory * price

        #calculating total value per supplier
        if inventory < 21:
            product_no_under_21[product_no] = inventory

        #sumtotal for product
        inventory_price.value  = inventory * price

    print(products_per_supplier)
    print(total_value_per_supplier)
    print(product_no_under_21)

    current_datetime = datetime.now()
    #inv_file.save("resources/new-inventory-with-sumtotal-"+current_datetime.strftime('%m-%d-%Y--%H-%M-%S')+".xlsx")

except Exception as e:
    print("Something happend")
    logging.error(e, exc_info=True)
    logger.error('Here comes the error.')

connect()

print("Last output from main.py file")

 